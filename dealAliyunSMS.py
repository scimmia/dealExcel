import xlrd
from openpyxl import load_workbook, Workbook
import threading
import json
import os
from shutil import copyfile
from tkinter import *
from tkinter.filedialog import askdirectory
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import *

def deal(templatePath, folderPath):
    showlog('初始化数据')
    bookTemp = xlrd.open_workbook(templatePath)
    sheetFrom = bookTemp.sheet_by_index(0)
    fileTo = os.path.join(folderPath, ('5阿里云-%d.xlsx' % (0)))
    f = open(fileTo,'w')
    f.close()
    wb = Workbook()
    sheet = wb.active
    sheet.title = '工作表1'
    # sheet = wb.create_sheet('工作表1')
    _ = sheet.cell(column=1, row=1, value='手机号')
    _ = sheet.cell(column=2, row=1, value='code')

    for rx in range(1,sheetFrom.nrows):
        rowNumber = rx % 5000 + 2
        if rowNumber == 2:
            index = rx / 5000
            wb.save(fileTo)
            fileTo = os.path.join(folderPath, ('5阿里云-%d.xlsx' % (index)))
            f = open(fileTo, 'w')
            f.close()
            wb = Workbook()
            sheet = wb.active
            sheet.title = '工作表1'
            _ = sheet.cell(column=1, row=1, value='手机号')
            _ = sheet.cell(column=2, row=1, value='code')
        row = (sheetFrom.row_values(rx))
        _ = sheet.cell(column=1, row=rowNumber, value=row[2])
        _ = sheet.cell(column=2, row=rowNumber, value=row[3])
    wb.save(fileTo)
    showlog('处理完成')


def doIt():
    folderPath = pathFolder.get()
    templatePath = pathTemplate.get()
    if len(folderPath) <= 0:
        showinfo('提示', '先选择文件夹')
    elif len(templatePath) <= 0:
        showinfo('提示', '选择模板文件')
    elif not templatePath.endswith('.xlsx'):
        showinfo('提示', '模板文件需先另存为xlsx格式')
    else:
        t.delete('1.0', 'end')
        threading.Thread(target=deal, args=(templatePath,folderPath,)).start()
        # deal(thePath)


def selectTemplate():
    path_ = askopenfilename()
    pathTemplate.set(path_)

def selectPath():
    path_ = askdirectory()
    pathFolder.set(path_)


def showlog(log):
    # print(log)
    t.insert('end', log + '\n')


def main():
    t.grid(row=3, columnspan=4)
    Entry(root, textvariable=pathTemplate).grid(row=0, column=0, columnspan=2)
    Button(root, text="选择模板文件", command=selectTemplate).grid(row=0, column=2)
    Button(root, text="开始合并", command=doIt).grid(row=0, column=3)
    Entry(root, textvariable=pathFolder).grid(row=1, column=0, columnspan=2)
    Button(root, text="选择文件夹   ", command=selectPath).grid(row=1, column=2)
    root.mainloop()


root = Tk()
pathTemplate = StringVar()
pathFolder = StringVar()
t = Text(root)
if __name__ == '__main__':
    main()
