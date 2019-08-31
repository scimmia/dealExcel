import xlrd
from openpyxl import load_workbook
import threading
import json
import os
from shutil import copyfile
from tkinter import *
from tkinter.filedialog import askdirectory
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import *

datasAll = {}


def checkSheet(sheetName):
    return sheetName.startswith('表')


def initDatas(filePath):
    global datasAll
    datasAll.clear()
    bookTemp = xlrd.open_workbook(filePath)
    for sheet in bookTemp.sheets():
        sheetName = sheet.name
        datas = {}
        if checkSheet(sheetName):
            for rx in range(sheet.nrows):
                row = (sheet.row_values(rx))
                if len(row) > 3 and isinstance(row[2], str) and row[2].endswith('农商银行'):
                    orgName = row[2]
                    datas[orgName] = {'rowNumber': rx}
            datasAll[sheetName] = datas
    # print(datas)
    # jsona = json.dumps(datasAll, ensure_ascii=False)
    # print(jsona)


def dealExcel(fileFrom):
    global datasAll
    bookTemp = xlrd.open_workbook(fileFrom)
    for sheetName in datasAll.keys():
        try:
            sheet = bookTemp.sheet_by_name(sheetName)
            datas = datasAll.get(sheetName)
            for rx in range(sheet.nrows):
                row = (sheet.row_values(rx))
                if len(row) > 3 and isinstance(row[2], str) and row[2].endswith('农商银行'):
                    orgName = row[2]
                    if not (orgName in datas):
                        showlog(orgName + ' 未找到')
                    else:
                        if sheet.cell_type(rx, 3) == 2 or sheet.cell_type(rx, 5) == 2 or sheet.cell_type(rx, 7) == 2:
                            if 'data' in datas[orgName]:
                                showlog('替换数据:\t' + sheetName + ':\t' + orgName)
                            datas[orgName]['data'] = row
        except:
            showlog(sheetName + ' 未找到')
            pass

    # jsona = json.dumps(datasAll, ensure_ascii=False)
    # print(jsona)


def writeExcel(fileTo):
    global datasAll
    wb = load_workbook(fileTo)
    for sheetName in datasAll.keys():
        datas = datasAll.get(sheetName)
        sheet = wb.get_sheet_by_name(sheetName)
        for orgName in datas.keys():
            rowNumber = datas[orgName]['rowNumber'] + 1
            if 'data' in datas[orgName]:
                data = datas[orgName]['data']
                try:
                    for col in range(3, len(data)):
                        theValue = data[col]
                        if isinstance(theValue, (float, int)):
                            theValue = int(theValue)
                        elif isinstance(theValue, str):
                            if not theValue.isdigit():
                                theValue = 0
                        _ = sheet.cell(column=col + 1, row=rowNumber, value=theValue)
                except:
                    showlog('写入错误：' + orgName)
                    pass
    wb.save(fileTo)


def listdir(path):  # 传入存储的list
    list_name = []
    for file in os.listdir(path):
        if (file.endswith('.xls') or file.endswith('.xlsx')) and (not file.startswith('all')):
            file_path = os.path.join(path, file)
            list_name.append(file_path)
    return list_name


def deal(templatePath, folderPath):
    fileAllName = '汇总.xlsx'
    showlog('初始化数据')
    initDatas(templatePath)

    for file in os.listdir(folderPath):
        if (file.endswith('.xls') or file.endswith('.xlsx')) and (not file.startswith(fileAllName)):
            showlog('开始处理:\t' + file)
            dealExcel(os.path.join(folderPath, file))

    showlog('写入到汇总文件')
    fileAll = os.path.join(folderPath, fileAllName)
    copyfile(templatePath, fileAll)
    writeExcel(fileAll)
    showlog('处理完成')
    showlog('汇总文件地址为：' + fileAll)
    showinfo('提示', '处理完成\n汇总文件地址为：\n' + fileAll)


def doIt():
    folderPath = pathFolder.get()
    templatePath = pathTemplate.get()
    if len(folderPath) <= 0:
        showinfo('提示', '先选择文件夹')
    elif len(templatePath) <= 0:
        showinfo('提示', '选择模板文件')
    elif not templatePath.endswith('.xlsx'):
        showinfo('提示', '模板文件需先另存为xlsx格式')
    else:
        t.delete('1.0', 'end')
        threading.Thread(target=deal, args=(templatePath,folderPath,)).start()
        # deal(thePath)


def selectTemplate():
    path_ = askopenfilename()
    pathTemplate.set(path_)

def selectPath():
    path_ = askdirectory()
    pathFolder.set(path_)


def showlog(log):
    # print(log)
    t.insert('end', log + '\n')


def main():
    t.grid(row=3, columnspan=4)
    Entry(root, textvariable=pathTemplate).grid(row=0, column=0, columnspan=2)
    Button(root, text="选择模板文件", command=selectTemplate).grid(row=0, column=2)
    Button(root, text="开始合并", command=doIt).grid(row=0, column=3)
    Entry(root, textvariable=pathFolder).grid(row=1, column=0, columnspan=2)
    Button(root, text="选择文件夹   ", command=selectPath).grid(row=1, column=2)
    root.mainloop()


root = Tk()
pathTemplate = StringVar()
pathFolder = StringVar()
t = Text(root)
if __name__ == '__main__':
    main()
