import xlrd
from openpyxl import load_workbook
import threading
import json
import os
from shutil import copyfile
from tkinter import *
from tkinter.filedialog import askdirectory
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import *

datasAll = {}


def checkSheet(sheetName):
    return sheetName.startswith('表')


def dealExcel(fileFrom):
    global datasAll
    bookTemp = xlrd.open_workbook(fileFrom)
    for sheetName in datasAll.keys():
        try:
            sheet = bookTemp.sheet_by_name(sheetName)
            datas = datasAll.get(sheetName)
            for rx in range(sheet.nrows):
                row = (sheet.row_values(rx))
                if len(row) > 3 and isinstance(row[2], str) and row[2].endswith('农商银行'):
                    orgName = row[2]
                    if not (orgName in datas):
                        showlog(orgName + ' 未找到')
                    else:
                        if sheet.cell_type(rx, 3) == 2 or sheet.cell_type(rx, 5) == 2 or sheet.cell_type(rx, 7) == 2:
                            if 'data' in datas[orgName]:
                                showlog('替换数据:\t' + sheetName + ':\t' + orgName)
                            datas[orgName]['data'] = row
        except:
            showlog(sheetName + ' 未找到')
            pass

    # jsona = json.dumps(datasAll, ensure_ascii=False)
    # print(jsona)


def writeExcel(fileTo):
    global datasAll
    wb = load_workbook(fileTo)
    for sheetName in datasAll.keys():
        datas = datasAll.get(sheetName)
        sheet = wb.get_sheet_by_name(sheetName)
        for orgName in datas.keys():
            rowNumber = datas[orgName]['rowNumber'] + 1
            if 'data' in datas[orgName]:
                data = datas[orgName]['data']
                try:
                    for col in range(3, len(data)):
                        theValue = data[col]
                        if isinstance(theValue, (float, int)):
                            theValue = int(theValue)
                        elif isinstance(theValue, str):
                            if not theValue.isdigit():
                                theValue = 0
                        _ = sheet.cell(column=col + 1, row=rowNumber, value=theValue)
                except:
                    showlog('写入错误：' + orgName)
                    pass
    wb.save(fileTo)


def listdir(path):  # 传入存储的list
    list_name = []
    for file in os.listdir(path):
        if (file.endswith('.xls') or file.endswith('.xlsx')) and (not file.startswith('all')):
            file_path = os.path.join(path, file)
            list_name.append(file_path)
    return list_name


def deal(filePath):
    showlog('初始化数据')
    bookTemp = xlrd.open_workbook(filePath)
    sheet = bookTemp.sheet_by_index(0)
    for rx in range(sheet.nrows):
        row = (sheet.row_values(rx))
        if isinstance(row[0], str) and row[0].endswith('题'):
            showlog(str(rx) + ' '+ row[1] + '  '+row[7])
            if not row[0].endswith('判断题'):
                if len(row[3]) + len(row[4]) > 26:
                    showlog('A. ' + row[3])
                    showlog('B. ' + row[4])
                else:
                    showlog('A. ' + row[3] + '    B. ' + row[4])
                if len(row[5]) + len(row[6]) > 26:
                    showlog('C. ' + row[5])
                    showlog('D. ' + row[6])
                else:
                    showlog('C. ' + row[5] + '    D. ' + row[6])
    # m = ''
    # for rx in range(sheet.nrows):
    #     row = (sheet.row_values(rx))
    #     if isinstance(row[0], str) and row[0].endswith('题'):
    #         m = m + str(rx) + '          '
    #         if rx % 5 ==0:
    #             m=m+'\n'
    # showlog(m)
    showinfo('提示', '处理完成\n')


def doIt():
    templatePath = pathTemplate.get()
    if len(templatePath) <= 0:
        showinfo('提示', '选择模板文件')
    elif not templatePath.endswith('.xlsx'):
        showinfo('提示', '模板文件需先另存为xlsx格式')
    else:
        t.delete('1.0', 'end')
        threading.Thread(target=deal, args=(templatePath,)).start()
        # deal(thePath)


def selectTemplate():
    path_ = askopenfilename()
    pathTemplate.set(path_)


def showlog(log):
    # print(log)
    t.insert('end', log + '\n')


def main():
    t.grid(row=3, columnspan=4)
    Entry(root, textvariable=pathTemplate).grid(row=0, column=0, columnspan=2)
    Button(root, text="选择文件", command=selectTemplate).grid(row=0, column=2)
    Button(root, text="开始", command=doIt).grid(row=0, column=3)
    root.mainloop()


root = Tk()
pathTemplate = StringVar()
pathFolder = StringVar()
t = Text(root)
if __name__ == '__main__':
    main()
