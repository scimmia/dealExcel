import xlrd
import xlwings as xw
from openpyxl import load_workbook

import json
import os
from shutil import copyfile


def initDatas(filePath):
    datas = {}
    bookTemp = xlrd.open_workbook(filePath)
    for sheet in bookTemp.sheets():
        sheetName = sheet.name
        if sheetName.startswith('表'):
            data = {}
            for rx in range(sheet.nrows):
                row = (sheet.row_values(rx))
                if len(row) > 3 and row[2].endswith('农商银行'):
                    data[row[2]] = rx
            datas[sheetName] = data
    # print(datas)
    jsona = json.dumps(datas, ensure_ascii=False)
    print(jsona)
    return datas


def dealExcel(fileFrom):
    datas = {}
    bookTemp = xlrd.open_workbook(fileFrom)
    for sheet in bookTemp.sheets():
        sheetName = sheet.name
        if sheetName.startswith('表'):
            data = []
            for rx in range(sheet.nrows):
                row = (sheet.row_values(rx))
                if len(row) > 3 and row[2].endswith('农商银行'):
                    if sheet.cell_type(rx,3) == 2 or sheet.cell_type(rx,5) == 2 or sheet.cell_type(rx,7) == 2:
                        data.append(row)
            datas[sheetName] = data
    # print(datas)
    # jsona = json.dumps(datas, ensure_ascii=False)
    # print(jsona)
    return datas


# def writeExcel(fileTo, datas, datasFromFile):
#     app = xw.App(visible=False, add_book=False)
#     wb = app.books.open(fileTo)
#     # wb = xw.Book(fileTo)
#
#     for key in datas.keys():
#         value = datas.get(key)
#         dataInserts = datasFromFile.get(key)
#         sheet = wb.sheets[key]
#         for dataInsert in dataInserts:
#             org = dataInsert[2]
#             rowNumber = value.get(org)
#             print(sheet.range((rowNumber, 3)).value)
#             # sheet.range((rowNumber, 3)).value = 'a'
#         # print(dataInserts)
#
#     wb.save()
#     wb.close()
#     # app.quit()
#     return datas
def writeExcel(fileTo, datas, datasFromFile):
    wb = load_workbook(fileTo)

    for key in datas.keys():
        value = datas.get(key)
        dataInserts = datasFromFile.get(key)
        sheet = wb.get_sheet_by_name(key)
        for dataInsert in dataInserts:
            org = dataInsert[2]
            rowNumber = value.get(org) + 1
            # print(rowNumber)

            # print(sheet.range((rowNumber, 3)).value)
            for col in range(3, len(dataInsert)):
                theValue = dataInsert[col]
                if isinstance (dataInsert[col],(float,int)):
                    theValue = int(dataInsert[col])
                # theValue = dataInsert[col]
                _ = sheet.cell(column=col+1, row=rowNumber, value= theValue)
            # sheet.cell(row=rowNumber, column=3).value = 'a'
        # print(dataInserts)

    wb.save(fileTo)
    return datas
# datas = initDatas()
# datasFromFile = dealExcel(r'H:\temp\excels\9.招远-6.30数据统计表（共10个表） - 副本.xls')
# writeExcel(r'H:\temp\all.xlsx',datas,datasFromFile)



def listdir(path):  # 传入存储的list
    list_name = []
    for file in os.listdir(path):
        if file.endswith('.xls') or file.endswith('.xlsx'):
            file_path = os.path.join(path, file)
            list_name.append(file_path)
    return list_name

def file_name(file_dir):
    for root, dirs, files in os.walk(file_dir):
        print(root)  # 当前目录路径
        print(dirs)  # 当前路径下所有子目录
        print(files)  # 当前路径下所有非目录子文件


def deal(fileFolder):
    datas = initDatas('./template.xlsx')
    fileAll = fileFolder+'\\all1.xlsx'
    copyfile('./template.xlsx', fileAll)
    for file in listdir(fileFolder):
        print('deal start', file)
        datasFromFile = dealExcel(file)
        writeExcel(fileAll, datas, datasFromFile)

deal(r'H:\temp\excels1')